# -*- coding: utf-8 -*-
"""
Ziraat Bankası ek tablo xlsx şablonu üretici.

Kaynak: KULLANICININ placeholder'ları elle yerleştirdiği xlsx (spec dosyası).
Bu script o dosyayı tarar, {{TOKEN}} içeren hücreleri bulur, TOKEN_MAP ile
app alanlarına eşler ve şunları üretir:

  templates/ziraat-ek-tablo.xlsx        — aynı dosya, STORED (sıkıştırmasız)
                                          olarak yeniden paketlenmiş.
  src/exports/ziraat-ek-tablo-manifest.json
                                        — {sheetIndex, sheet, cell, token,
                                          field, type} listesi; tarayıcıdaki
                                          doldurma motoru bunu kullanır.

ÖNEMLİ: Şablon STORED olmalıdır. Tarayıcı doldurma motoru (src/exports/
xlsx-fill.js) inflate/deflate kütüphanesi kullanmaz; zip girişlerini ham
okur. Excel'de açıp kaydederseniz dosya DEFLATE'e döner — bu script'i
yeniden çalıştırıp STORED'a çevirin.

Kullanım:
    python tools/build-ziraat-ek-tablo-xlsx.py [kaynak.xlsx]
"""
import json
import os
import re
import sys
import zipfile

import openpyxl

DEFAULT_SRC = r"C:\Users\90551\OneDrive\Masaüstü\ziraat-ek-tablo.xlsx"
SRC = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
APP = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_XLSX = os.path.join(APP, "templates", "ziraat-ek-tablo.xlsx")
OUT_MANIFEST = os.path.join(APP, "src", "exports", "ziraat-ek-tablo-manifest.json")

TOKEN_RE = re.compile(r"\{\{([^}]+)\}\}")


def fold(name):
    """Türkçe-katlama + noktalama temizleme (template-engine foldTokenName ile aynı ruh)."""
    out = str(name or "")
    for a, b in (("İ", "I"), ("ı", "i"), ("Ç", "C"), ("ç", "c"), ("Ğ", "G"), ("ğ", "g"),
                 ("Ö", "O"), ("ö", "o"), ("Ş", "S"), ("ş", "s"), ("Ü", "U"), ("ü", "u")):
        out = out.replace(a, b)
    return re.sub(r"[^A-Z0-9]+", "", out.upper())


# TOKEN (katlanmış) -> (app alan anahtarı | özel çözümleyici, tip)
# type: "number" → hücreye sayı yazılır (formüller çalışır)
#       "text"   → hücreye metin yazılır
TOKEN_MAP = {
    # --- tapu / arsa ---
    fold("ADA"): ("blockNo", "text"),
    fold("PARSEL"): ("parcelNo", "text"),
    fold("LAND_AREA"): ("landArea", "number"),
    fold("POST_ROAD_SETBACK_PARCEL_AREA"): ("ZRT_POST_SETBACK_AREA", "number"),
    # --- imar ---
    fold("LEGEND"): ("legend", "text"),
    fold("ORDER"): ("order", "text"),
    fold("CALCULATED_EMSAL"): ("calculatedEmsal", "number"),
    # --- gayrimenkul tanımı ---
    fold("GAYRIMENKUL_ADI"): ("titleQuality", "text"),
    fold("YAPI_SINIFI"): ("buildingClass", "text"),
    fold("MAİN_PROPERTY_FLOOR_COUNT_TEXT"): ("mainPropertyFloorCountText", "text"),
    # Belge: iskan varsa iskan, yoksa en güncel ruhsat (özel çözümleyici).
    fold("YAPI_BELGESI"): ("ZRT_BUILDING_DOC", "text"),
    # --- alanlar ---
    fold("TOTAL_LEGAL_AREA"): ("legalValueArea", "number"),
    fold("TOTAL_CURRENT_AREA"): ("currentValueArea", "number"),
    # --- değerler ---
    fold("LEGAL_VALUE"): ("legalValue", "number"),
    fold("CURRENT_VALUE"): ("currentValue", "number"),
    fold("LAND_VALUE"): ("landValue", "number"),
    fold("LEGAL_BUİLDİNG_VALUE"): ("legalBuildingValue", "number"),
    fold("CURRENT_BUİLDİNG_VALUE"): ("currentBuildingValue", "number"),
    fold("LEGAL_PREMİUM_VALUE"): ("legalPremiumValue", "number"),
    fold("CURRENT_PREMİUM_VALUE"): ("currentPremiumValue", "number"),
    # --- diğer ---
    fold("SALEABİLİTY"): ("saleability", "text"),
}


# Kaynak dosyada {{TOKEN}} konulmamış ama uygulamanın doldurması istenen
# hücreler (kullanıcı onayı 2026-07-21). Sayı hücreleri olduklarından şablonda
# 0 dururlar; manifest üzerinden doldurulurlar.
EXTRA_CELLS = {
    "ARSA": [
        # Kullanıcı kuralı: yapı nizamı BİTİŞİK ise KAKS yerine kat adedi yazılır
        # (app'teki composeImarCalculatedEmsal mantığının aynısı).
        ("F3", "ZRT_KAKS_OR_FLOOR", "number"),
        ("G3", "taks", "number"),
        ("H3", "hmax", "number"),
    ],
    "KONUT-İŞYERLERİ": [
        ("G3", "legalValue", "number"),
        ("G10", "currentValue", "number"),
    ],
}

# Formül hücrelerinin ÖNBELLEK (<v>) değerleri. openpyxl formülleri cache'siz
# yazar; Excel açılışta hesaplar ama cache okuyan araçlar boş görür. Bu yüzden
# doldurma sırasında formülü koruyup <v> önbelleğini de yazarız.
# type: formulaText | formulaNumber, field: JS tarafındaki çözümleyici/alan.
FORMULA_CACHE = {
    "KONUT-İŞYERLERİ": [
        ("C10", "titleQuality", "formulaText"),        # =C3
        ("D10", "ZRT_BUILDING_DOC", "formulaText"),    # =D3
        ("F3", "ZRT_LEGAL_UNIT_VALUE", "formulaNumber"),    # =IFERROR(G3/E3,0)
        ("F10", "ZRT_CURRENT_UNIT_VALUE", "formulaNumber"), # =IFERROR(G10/E10,0)
        ("E6", "legalValueArea", "formulaNumber"),     # =SUM(E3:E5)
        ("G6", "legalValue", "formulaNumber"),         # =SUM(G3:G5)
        ("E14", "currentValueArea", "formulaNumber"),  # =SUM(E10:E13)
        ("G14", "currentValue", "formulaNumber"),      # =SUM(G10:G13)
    ],
}


# Kullanıcı isteği (2026-07-21): yüzölçümü, m² birim değeri, Hmax ve terk
# sonrası parsel büyüklüğü hücrelerinde virgülden sonra 2 basamak gösterilmeli
# (ör. 193 yerine 192,74). Sayı biçimi "#,##0.00" olarak zorlanır.
TWO_DECIMAL_CELLS = {
    "TARLA": ["D3", "E3", "D8", "D12", "E12", "D17"],  # yüzölçüm + m² birim değer + toplamlar
    "ARSA": ["D3", "H3", "J3", "D8", "D12", "E12", "D17", "D21", "E21", "D26"],  # yüzölçüm, Hmax, terk sonrası, m² birim
    "KONUT-İŞYERLERİ": ["E3", "F3", "E6", "E10", "F10", "E14"],  # alan + m² birim değer
    "NİTELİKLİ GAYRİMENKUL": ["G3", "H3", "G4", "H4", "G9", "G13", "H13", "G14", "G18"],  # alan + birim değer + toplam
}

# Formül override'ları (kullanıcı isteği): NİTELİKLİ GAYRİMENKUL toplam
# hücreleri, doğru satır aralığına işaret etmeli.
FORMULA_OVERRIDES = {
    "NİTELİKLİ GAYRİMENKUL": [
        ("G9", "=SUM(G4:G8)"),    # =TOPLA(G4:G8)
        ("G18", "=SUM(G14:G17)"),  # =TOPLA(G14:G17)
    ],
}


def main():
    if not os.path.exists(SRC):
        raise SystemExit(f"Kaynak bulunamadi: {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=False)
    sheet_order = wb.sheetnames
    manifest = []
    unknown = []

    # Formül override'ları (toplam aralıkları vb.).
    for sheet_name, overrides in FORMULA_OVERRIDES.items():
        ws = wb[sheet_name]
        for coord, formula in overrides:
            ws[coord] = formula

    # 2 ondalık basamak biçimi (yüzölçümü, m² birim değer, Hmax, terk sonrası).
    for sheet_name, coords in TWO_DECIMAL_CELLS.items():
        ws = wb[sheet_name]
        for coord in coords:
            ws[coord].number_format = "#,##0.00"

    for ws in wb.worksheets:
        sheet_index = sheet_order.index(ws.title) + 1  # sheet1.xml = 1
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                tokens = TOKEN_RE.findall(cell.value)
                if not tokens:
                    continue
                if len(tokens) > 1:
                    unknown.append(f"{ws.title}!{cell.coordinate}: tek hucrede birden fazla token ({tokens})")
                    continue
                token = tokens[0].strip()
                spec = TOKEN_MAP.get(fold(token))
                if not spec:
                    unknown.append(f"{ws.title}!{cell.coordinate}: bilinmeyen token {{{{{token}}}}}")
                    continue
                field, typ = spec
                manifest.append({
                    "sheetIndex": sheet_index,
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "token": "{{" + token + "}}",
                    "field": field,
                    "type": typ,
                })

    if unknown:
        raise SystemExit("Eslenemeyen token(lar):\n  " + "\n  ".join(unknown))

    # Token'sız ama doldurulacak hücreler + formül önbellek girişleri.
    for sheet_name, extras in EXTRA_CELLS.items():
        idx = sheet_order.index(sheet_name) + 1
        for coord, field, typ in extras:
            manifest.append({"sheetIndex": idx, "sheet": sheet_name, "cell": coord,
                             "token": "", "field": field, "type": typ})

    for sheet_name, caches in FORMULA_CACHE.items():
        ws = wb[sheet_name]
        idx = sheet_order.index(sheet_name) + 1
        for coord, field, typ in caches:
            value = ws[coord].value
            if not (isinstance(value, str) and value.startswith("=")):
                raise SystemExit(f"{sheet_name}!{coord} formul degil ({value!r}); FORMULA_CACHE guncellenmeli.")
            manifest.append({"sheetIndex": idx, "sheet": sheet_name, "cell": coord,
                             "token": "", "field": field, "type": typ})

    tmp = OUT_XLSX + ".deflate.tmp"
    wb.save(tmp)

    # STORED (sıkıştırmasız) yeniden paketle → tarayıcı kütüphanesiz okuyabilsin.
    with zipfile.ZipFile(tmp, "r") as zin:
        names = zin.namelist()
        with zipfile.ZipFile(OUT_XLSX, "w", compression=zipfile.ZIP_STORED) as zout:
            for n in names:
                zout.writestr(n, zin.read(n))
    os.remove(tmp)

    os.makedirs(os.path.dirname(OUT_MANIFEST), exist_ok=True)
    with open(OUT_MANIFEST, "w", encoding="utf-8") as f:
        json.dump({
            "template": "templates/ziraat-ek-tablo.xlsx",
            "sheets": sheet_order,
            "cells": manifest,
        }, f, ensure_ascii=False, indent=2)

    with zipfile.ZipFile(OUT_XLSX) as z:
        comps = set(i.compress_type for i in z.infolist())

    print("OK ->", OUT_XLSX)
    print("OK ->", OUT_MANIFEST)
    print("manifest hucre sayisi:", len(manifest))
    print("compress types (0=STORED):", comps)
    per_sheet = {}
    for m in manifest:
        per_sheet[m["sheet"]] = per_sheet.get(m["sheet"], 0) + 1
    for s in sheet_order:
        print(f"  {s}: {per_sheet.get(s, 0)} hucre")


if __name__ == "__main__":
    main()
